import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import cv2
from pyzbar.pyzbar import decode
import csv
import xml.etree.ElementTree as ET
import pandas as pd
from datetime import datetime
import os
import threading
from PIL import Image, ImageTk
import numpy as np
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from google.auth.transport.requests import Request
import pickle
import json
import openpyxl

class EnhancedQRScannerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Enhanced QR Scanner")
        self.root.geometry("800x600")
        self.root.protocol("WM_DELETE_WINDOW", self.safe_close)
        
        # Theme and styling
        style = ttk.Style()
        style.theme_use('clam')
        
        # Variables
        self.spreadsheet_path = None
        self.is_scanning = False
        self.current_camera = 0
        self.available_cameras = self.get_available_cameras()
        self.last_scan = None
        self.duplicate_check = tk.BooleanVar(value=True)
        self.auto_save = tk.BooleanVar(value=True)
        self.frozen_frame = None
        self.current_sl_no = 1
        self.google_sheets_enabled = tk.BooleanVar(value=False)
        self.sheets_service = None
        self.spreadsheet_id = None
        
        # Google Sheets API scope
        self.SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
        
        self.setup_ui()
        
    def setup_ui(self):
        # Main container
        self.main_container = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        self.main_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Left panel setup
        self.setup_left_panel()
        
        # Right panel setup
        self.setup_right_panel()
        
        # Scan confirmation dialog
        self.setup_confirmation_dialog()
        
        # Status bar
        self.status_bar = ttk.Label(self.root, text="Ready", relief=tk.SUNKEN)
        self.status_bar.pack(fill=tk.X, side=tk.BOTTOM, padx=5, pady=2)
        
    def setup_left_panel(self):
        self.left_panel = ttk.Frame(self.main_container)
        self.main_container.add(self.left_panel, weight=2)
        
        # Camera feed
        self.camera_frame = ttk.LabelFrame(self.left_panel, text="Camera Feed")
        self.camera_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.video_label = ttk.Label(self.camera_frame)
        self.video_label.pack(fill=tk.BOTH, expand=True)
        
        # Camera controls
        self.controls_frame = ttk.Frame(self.left_panel)
        self.controls_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.camera_combo = ttk.Combobox(self.controls_frame, 
                                       values=[f"Camera {i}" for i in self.available_cameras])
        self.camera_combo.set("Camera 0")
        self.camera_combo.pack(side=tk.LEFT, padx=5)
        
        self.scan_button = ttk.Button(self.controls_frame, 
                                    text="Start Scanning", 
                                    command=self.toggle_scanning)
        self.scan_button.pack(side=tk.LEFT, padx=5)
        
        self.close_button = ttk.Button(self.controls_frame,
                                     text="Safe Close",
                                     command=self.safe_close)
        self.close_button.pack(side=tk.RIGHT, padx=5)

    def setup_right_panel(self):
        self.right_panel = ttk.Frame(self.main_container)
        self.main_container.add(self.right_panel, weight=1)
        
        # Settings section
        self.settings_frame = ttk.LabelFrame(self.right_panel, text="Settings")
        self.settings_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Checkbutton(self.settings_frame, 
                       text="Check for duplicates", 
                       variable=self.duplicate_check).pack(anchor=tk.W, padx=5, pady=2)
        
        ttk.Checkbutton(self.settings_frame, 
                       text="Auto-save", 
                       variable=self.auto_save).pack(anchor=tk.W, padx=5, pady=2)
        
        ttk.Checkbutton(self.settings_frame,
                       text="Use Google Sheets",
                       variable=self.google_sheets_enabled,
                       command=self.toggle_google_sheets).pack(anchor=tk.W, padx=5, pady=2)
        
        # File format selection
        self.file_format = ttk.Combobox(self.settings_frame,
                                      values=["CSV", "XML", "Excel", "Google Sheets"])
        self.file_format.set("CSV")
        self.file_format.pack(anchor=tk.W, padx=5, pady=2)
        
        # Spreadsheet section
        self.spreadsheet_frame = ttk.LabelFrame(self.right_panel, text="Spreadsheet")
        self.spreadsheet_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.link_button = ttk.Button(self.spreadsheet_frame, 
                                    text="Link Spreadsheet", 
                                    command=self.link_spreadsheet)
        self.link_button.pack(fill=tk.X, padx=5, pady=5)
        
        self.path_label = ttk.Label(self.spreadsheet_frame, 
                                  text="No spreadsheet linked", 
                                  wraplength=200)
        self.path_label.pack(fill=tk.X, padx=5)
        
        # Scan history
        self.history_frame = ttk.LabelFrame(self.right_panel, text="Scan History")
        self.history_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.history_tree = ttk.Treeview(self.history_frame, 
                                       columns=("SL No.", "Time", "Data"),
                                       show="headings")
        self.history_tree.heading("SL No.", text="SL No.")
        self.history_tree.heading("Time", text="Time")
        self.history_tree.heading("Data", text="Data")
        self.history_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    def setup_confirmation_dialog(self):
        self.confirm_dialog = tk.Toplevel(self.root)
        self.confirm_dialog.withdraw()
        self.confirm_dialog.transient(self.root)
        self.confirm_dialog.title("Scan Confirmation")
        
        self.confirm_label = ttk.Label(self.confirm_dialog, text="", wraplength=300)
        self.confirm_label.pack(padx=20, pady=10)
        
        self.confirm_button = ttk.Button(self.confirm_dialog,
                                       text="OK",
                                       command=self.confirm_scan)
        self.confirm_button.pack(pady=10)

    def toggle_google_sheets(self):
        if self.google_sheets_enabled.get():
            self.setup_google_sheets()
        else:
            self.sheets_service = None
            self.spreadsheet_id = None

    def setup_google_sheets(self):
        creds = None
        if os.path.exists('token.pickle'):
            with open('token.pickle', 'rb') as token:
                creds = pickle.load(token)
        
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    'credentials.json', self.SCOPES)
                creds = flow.run_local_server(port=0)
            
            with open('token.pickle', 'wb') as token:
                pickle.dump(creds, token)
        
        try:
            self.sheets_service = build('sheets', 'v4', credentials=creds)
            messagebox.showinfo("Success", "Google Sheets connected successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to connect to Google Sheets: {str(e)}")
            self.google_sheets_enabled.set(False)

    def scan_qr(self):
        cap = cv2.VideoCapture(self.current_camera)
        if not cap.isOpened():
            messagebox.showerror("Error", "Failed to access the camera")
            self.is_scanning = False
            return
        
        while self.is_scanning:
            ret, frame = cap.read()
            if not ret:
                continue
            
            # Convert frame for display
            cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)
            img = Image.fromarray(cv2image)
            imgtk = ImageTk.PhotoImage(image=img)
            self.video_label.imgtk = imgtk
            self.video_label.configure(image=imgtk)
            
            # Scan for QR codes
            decoded_objects = decode(frame)
            for obj in decoded_objects:
                qr_data = obj.data.decode("utf-8")
                if self.process_scan(qr_data):
                    self.frozen_frame = frame.copy()
                    self.show_confirmation(qr_data)
                    cap.release()
                    return
            
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break
        
        cap.release()
        self.video_label.configure(image='')

    def show_confirmation(self, data):
        self.confirm_label.configure(text=f"Scanned Data:\n{data}")
        self.confirm_dialog.deiconify()
        self.confirm_dialog.geometry(f"+{self.root.winfo_x() + 50}+{self.root.winfo_y() + 50}")

    def confirm_scan(self):
        self.confirm_dialog.withdraw()
        self.last_scan = None
        self.frozen_frame = None
        self.is_scanning = True
        threading.Thread(target=self.scan_qr, daemon=True).start()

    def process_scan(self, data):
        """Process scanned QR code data"""
        current_time = datetime.now()
        
        # Check for duplicate if enabled
        if self.duplicate_check.get():
            if data == self.last_scan:
                self.update_status(f"Duplicate scan ignored: {data}")
                return False
        
        self.last_scan = data
        
        # Add to history
        self.history_tree.insert('', 0, values=(self.current_sl_no, 
                                              current_time.strftime('%H:%M:%S'), 
                                              data))
        
        # Save to file if auto-save is enabled
        if self.auto_save.get():
            self.save_data(data)
        
        self.current_sl_no += 1
        self.update_status(f"Scanned: {data}")
        return True

    def save_data(self, data):
        if not self.spreadsheet_path and not self.google_sheets_enabled.get():
            return False
        
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        if self.google_sheets_enabled.get():
            return self.save_to_google_sheets(current_time, data)
        
        file_format = self.file_format.get()
        try:
            if file_format == "CSV":
                self.save_to_csv(current_time, data)
            elif file_format == "XML":
                self.save_to_xml(current_time, data)
            elif file_format == "Excel":
                self.save_to_excel(current_time, data)
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save data: {str(e)}")
            return False

    def save_to_csv(self, timestamp, data):
        if not os.path.exists(self.spreadsheet_path):
            with open(self.spreadsheet_path, 'w', newline='') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(['SL No.', 'Timestamp', 'Data'])
        
        with open(self.spreadsheet_path, 'a', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow([self.current_sl_no, timestamp, data])

    def save_to_xml(self, timestamp, data):
        if not os.path.exists(self.spreadsheet_path):
            root = ET.Element("scans")
            tree = ET.ElementTree(root)
        else:
            tree = ET.parse(self.spreadsheet_path)
            root = tree.getroot()
        
        scan = ET.SubElement(root, "scan")
        ET.SubElement(scan, "sl_no").text = str(self.current_sl_no)
        ET.SubElement(scan, "timestamp").text = timestamp
        ET.SubElement(scan, "data").text = data
        
        tree.write(self.spreadsheet_path)

    def save_to_excel(self, timestamp, data):
        if not os.path.exists(self.spreadsheet_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(['SL No.', 'Timestamp', 'Data'])
        else:
            wb = openpyxl.load_workbook(self.spreadsheet_path)
            ws = wb.active
        
        ws.append([self.current_sl_no, timestamp, data])
        wb.save(self.spreadsheet_path)

    def save_to_google_sheets(self, timestamp, data):
        if not self.sheets_service or not self.spreadsheet_id:
            return False
        
        try:
            values = [[self.current_sl_no, timestamp, data]]
            body = {'values': values}
            self.sheets_service.spreadsheets().values().append(
                spreadsheetId=self.spreadsheet_id,
                range='A:C',
                valueInputOption='RAW',
                insertDataOption='INSERT_ROWS',
                body=body).execute()
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save to Google Sheets: {str(e)}")
            return False

    def setup_google_sheets(self):
        """Setup Google Sheets authentication using direct OAuth flow"""
        # Client configuration for installed applications
        CLIENT_CONFIG = {
            "installed": {
                "client_id": "YOUR_CLIENT_ID",  # Replace with your OAuth client ID
                "project_id": "quickstart-1234",
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
                "client_secret": "YOUR_CLIENT_SECRET",  # Replace with your OAuth client secret
                "redirect_uris": ["http://localhost", "urn:ietf:wg:oauth:2.0:oob"]
            }
        }

        def create_oauth_dialog():
            """Create a dialog to input OAuth credentials"""
            dialog = tk.Toplevel(self.root)
            dialog.title("Google OAuth Setup")
            dialog.geometry("600x400")
            dialog.transient(self.root)
            dialog.grab_set()

            # Instructions
            ttk.Label(dialog, text="To enable Google Sheets integration, you need to:", 
                    wraplength=550).pack(padx=20, pady=10)
            
            instructions = [
                "1. Go to https://console.cloud.google.com",
                "2. Create a new project or select existing one",
                "3. Enable the Google Sheets API",
                "4. Go to 'Credentials' and click 'Create Credentials'",
                "5. Choose 'OAuth client ID'",
                "6. Select 'Desktop app' as application type",
                "7. Copy the provided Client ID and Client Secret below"
            ]
            
            for instruction in instructions:
                ttk.Label(dialog, text=instruction, wraplength=550).pack(padx=20, pady=2)

            # Create entry fields
            client_id_frame = ttk.Frame(dialog)
            client_id_frame.pack(fill=tk.X, padx=20, pady=5)
            ttk.Label(client_id_frame, text="Client ID:").pack(side=tk.LEFT)
            client_id_entry = ttk.Entry(client_id_frame, width=50)
            client_id_entry.pack(side=tk.LEFT, padx=5)

            client_secret_frame = ttk.Frame(dialog)
            client_secret_frame.pack(fill=tk.X, padx=20, pady=5)
            ttk.Label(client_secret_frame, text="Client Secret:").pack(side=tk.LEFT)
            client_secret_entry = ttk.Entry(client_secret_frame, width=50)
            client_secret_entry.pack(side=tk.LEFT, padx=5)

            def validate_and_save():
                client_id = client_id_entry.get().strip()
                client_secret = client_secret_entry.get().strip()
                
                if not client_id or not client_secret:
                    messagebox.showerror("Error", "Please enter both Client ID and Client Secret")
                    return
                
                # Update client config
                CLIENT_CONFIG["installed"]["client_id"] = client_id
                CLIENT_CONFIG["installed"]["client_secret"] = client_secret
                
                dialog.destroy()
                self.complete_oauth_flow(CLIENT_CONFIG)

            ttk.Button(dialog, text="Continue", command=validate_and_save).pack(pady=20)
            
            # Add hyperlink to open browser
            def open_console():
                import webbrowser
                webbrowser.open("https://console.cloud.google.com")
            
            link = ttk.Label(dialog, text="Open Google Cloud Console", 
                            foreground="blue", cursor="hand2")
            link.pack(pady=10)
            link.bind("<Button-1>", lambda e: open_console())

        def get_saved_credentials():
            """Check for saved credentials"""
            if os.path.exists('token.pickle'):
                with open('token.pickle', 'rb') as token:
                    creds = pickle.load(token)
                    if creds and creds.valid:
                        return creds
                    if creds and creds.expired and creds.refresh_token:
                        try:
                            creds.refresh(Request())
                            return creds
                        except Exception:
                            os.remove('token.pickle')
            return None

        def save_credentials(creds):
            """Save credentials to file"""
            with open('token.pickle', 'wb') as token:
                pickle.dump(creds, token)

        # First check for saved credentials
        creds = get_saved_credentials()
        if creds:
            try:
                self.sheets_service = build('sheets', 'v4', credentials=creds)
                self.get_spreadsheet_link()
                return
            except Exception:
                os.remove('token.pickle')
        
        # If no valid credentials, show OAuth dialog
        create_oauth_dialog()

    def complete_oauth_flow(self, client_config):
        """Complete the OAuth flow after getting client credentials"""
        try:
            flow = InstalledAppFlow.from_client_config(
                client_config, self.SCOPES,
                redirect_uri='urn:ietf:wg:oauth:2.0:oob')
            
            # Open the authorization URL in the default browser
            auth_url = flow.authorization_url()[0]
            import webbrowser
            webbrowser.open(auth_url)
            
            # Create authorization code dialog
            auth_dialog = tk.Toplevel(self.root)
            auth_dialog.title("Enter Authorization Code")
            auth_dialog.geometry("500x200")
            auth_dialog.transient(self.root)
            auth_dialog.grab_set()
            
            ttk.Label(auth_dialog, 
                    text="Please copy the authorization code from the browser:",
                    wraplength=450).pack(padx=20, pady=10)
            
            code_entry = ttk.Entry(auth_dialog, width=50)
            code_entry.pack(padx=20, pady=10)
            
            def complete_auth():
                
                code = code_entry.get().strip()
                if not code:
                    messagebox.showerror("Error", "Please enter the authorization code")
                    return
                
                try:
                    flow.fetch_token(code=code)
                    creds = flow.credentials
                    
                    # Save credentials
                    with open('token.pickle', 'wb') as token:
                        pickle.dump(creds, token)
                    
                    self.sheets_service = build('sheets', 'v4', credentials=creds)
                    auth_dialog.destroy()
                    self.get_spreadsheet_link()
                    
                except Exception as e:
                    messagebox.showerror("Error", 
                        f"Failed to authenticate: {str(e)}")
                    auth_dialog.destroy()
                    self.google_sheets_enabled.set(False)
            
            ttk.Button(auth_dialog, text="Submit", 
                    command=complete_auth).pack(pady=20)


    def get_spreadsheet_link(self):
        """Prompt user for Google Sheets link and extract spreadsheet ID"""
        spreadsheet_link = simpledialog.askstring(
            "Google Sheets", 
            "Enter the Google Sheets URL:\n(Make sure the sheet is shared with edit access)",
            parent=self.root
        )
        
        if not spreadsheet_link:
            self.google_sheets_enabled.set(False)
            return
        
        try:
            # Extract spreadsheet ID from the URL
            if "/d/" in spreadsheet_link:
                self.spreadsheet_id = spreadsheet_link.split("/d/")[1].split("/")[0]
            else:
                raise ValueError("Invalid Google Sheets URL")
            
            # Verify access to the spreadsheet
            try:
                self.sheets_service.spreadsheets().get(
                    spreadsheetId=self.spreadsheet_id).execute()
                
                # Check if headers exist, if not add them
                result = self.sheets_service.spreadsheets().values().get(
                    spreadsheetId=self.spreadsheet_id,
                    range='A1:C1').execute()
                
                if 'values' not in result or result['values'][0] != ['SL No.', 'Timestamp', 'Data']:
                    values = [['SL No.', 'Timestamp', 'Data']]
                    body = {'values': values}
                    self.sheets_service.spreadsheets().values().update(
                        spreadsheetId=self.spreadsheet_id,
                        range='A1:C1',
                        valueInputOption='RAW',
                        body=body).execute()
                
                # Get the sheet title for display
                sheet_metadata = self.sheets_service.spreadsheets().get(
                    spreadsheetId=self.spreadsheet_id).execute()
                sheet_title = sheet_metadata['properties']['title']
                
                self.path_label.config(text=f"Linked to Google Sheet: {sheet_title}")
                messagebox.showinfo("Success", 
                    "Successfully connected to Google Sheet!")
                
            except Exception as e:
                messagebox.showerror("Error", 
                    "Unable to access the spreadsheet. Please check the URL and sharing permissions.")
                self.google_sheets_enabled.set(False)
                self.spreadsheet_id = None
                
        except Exception as e:
            messagebox.showerror("Error", 
                f"Invalid Google Sheets URL: {str(e)}")
            self.google_sheets_enabled.set(False)
            self.spreadsheet_id = None

    def link_spreadsheet(self):
        """Link a spreadsheet file or Google Sheet for saving scan data"""
        if self.google_sheets_enabled.get():
            self.setup_google_sheets()
        else:
            # File format selection
            file_format = self.file_format.get()
            file_types = []
            
            if file_format == "CSV":
                file_types = [("CSV files", "*.csv")]
                default_ext = ".csv"
            elif file_format == "XML":
                file_types = [("XML files", "*.xml")]
                default_ext = ".xml"
            elif file_format == "Excel":
                file_types = [("Excel files", "*.xlsx")]
                default_ext = ".xlsx"
            
            filepath = filedialog.asksaveasfilename(
                defaultextension=default_ext,
                filetypes=file_types
            )
            
            if filepath:
                self.spreadsheet_path = filepath
                self.path_label.config(text=f"Linked to: {os.path.basename(filepath)}")
                
                # Initialize the file with headers based on format
                if file_format == "CSV":
                    with open(filepath, 'w', newline='') as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow(['SL No.', 'Timestamp', 'Data'])
                elif file_format == "XML":
                    root = ET.Element("scans")
                    tree = ET.ElementTree(root)
                    tree.write(filepath)
                elif file_format == "Excel":
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.append(['SL No.', 'Timestamp', 'Data'])
                    wb.save(filepath)
                        
    def get_available_cameras(self):
        """Get list of available camera indices"""
        available_cameras = []
        for i in range(10):  # Check first 10 camera indices
            cap = cv2.VideoCapture(i)
            if cap.isOpened():
                available_cameras.append(i)
                cap.release()
        return available_cameras

    def toggle_scanning(self):
        """Toggle QR code scanning on/off"""
        if not self.is_scanning:
            if not self.spreadsheet_path and not (self.google_sheets_enabled.get() and self.spreadsheet_id):
                messagebox.showwarning("Warning", "Please link a spreadsheet first!")
                return
            
            self.is_scanning = True
            self.scan_button.config(text="Stop Scanning")
            self.current_camera = int(self.camera_combo.get().split()[-1])
            threading.Thread(target=self.scan_qr, daemon=True).start()
        else:
            self.is_scanning = False
            self.scan_button.config(text="Start Scanning")
    
    def update_status(self, message):
        """Update status bar message"""
        self.status_bar.config(text=message)
    
    def safe_close(self):
        """Safely close the application"""
        if messagebox.askyesno("Quit", "Are you sure you want to quit?"):
            self.is_scanning = False
            self.root.after(100, self.root.destroy)

def main():
    root = tk.Tk()
    app = EnhancedQRScannerApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
